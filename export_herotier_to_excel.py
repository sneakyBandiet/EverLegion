import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.filters import AutoFilter

# Color mapping based on your images
COLOR_MAP = {
    "S": "FF0000",    # Red
    "S+": "C00000",   # Dark Red
    "A": "FFFF00",    # Yellow
    "B": "00B0F0",    # Blue
    "C": "00FF00",    # Green
    "D": "92D050",    # Light Green
    "F": "0070C0"     # Dark Blue
}

# Columns to color
COLOR_COLUMNS = ["PvP", "Story/Tower", "Faction Tower", "Bosses"]

# Load JSON data
with open("herotier_full.json", "r") as f:
    data = json.load(f)

wb = Workbook()
ws = wb.active
ws.title = "Hero Tier"

# Write header
headers = list(data[0].keys())

import json
import sys
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def export_json_to_excel(json_path, excel_path):
    with open(json_path, 'r') as f:
        data = json.load(f)

    wb = Workbook()
    ws = wb.active
    ws.title = "Hero Tier List"

    headers = list(data[0].keys())
    ws.append(headers)

    color_map = {
        'S+': 'FFD700', # Gold
        'S': 'C0C0C0',  # Silver
        'A': '00FF00',  # Green
        'B': '00BFFF',  # Blue
        'C': 'FF8C00',  # Dark Orange
        'D': 'FF69B4',  # Pink
        'F': 'FF0000',  # Red
        '-': 'FFFFFF',  # White
    }

    for row in data:
        ws.append([row.get(h, '') for h in headers])

    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            val = str(cell.value)
            if val in color_map:
                cell.fill = PatternFill(start_color=color_map[val], end_color=color_map[val], fill_type="solid")

    ws.auto_filter.ref = ws.dimensions
    wb.save(excel_path)

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python export_herotier_to_excel.py <input_json> <output_excel>")
        sys.exit(1)
    export_json_to_excel(sys.argv[1], sys.argv[2])
