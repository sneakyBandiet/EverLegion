import json
import os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill

OUT_DIR = "out"
EXCEL_PATH = os.path.join(OUT_DIR, "herotier_full.xlsx")

COLOR_MAP = {
    'S+': '8B0000', # Dark Red
    'S': 'FF0000',  # Red
    'A': 'FFA500',  # Orange
    'B': 'FFFF00',  # Yellow
    'C': '90EE90',  # Light Green
    'D': '00FF00',  # Green
    'F': '00BFFF',  # Blue
    '-': 'FFFFFF',  # White
}

INVALID_TITLE_CHARS = set(r'\/:*?[]')

def safe_sheet_title(title: str, existing: set) -> str:
    cleaned = ''.join(ch for ch in str(title) if ch not in INVALID_TITLE_CHARS).strip()[:31] or "Sheet"
    base, i = cleaned, 2
    while cleaned in existing:
        suffix = f" ({i})"
        cleaned = (base[:31-len(suffix)] + suffix)
        i += 1
    existing.add(cleaned)
    return cleaned

def apply_colors(ws):
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            val = str(cell.value)
            if val in COLOR_MAP:
                cell.fill = PatternFill(start_color=COLOR_MAP[val],
                                        end_color=COLOR_MAP[val],
                                        fill_type="solid")
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

def export_json_to_excel(json_path, excel_path):
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    if not data:
        raise ValueError("Input JSON has no rows.")

    wb = Workbook()
    ws_full = wb.active
    ws_full.title = "Hero Tier List"

    headers = list(data[0].keys())
    ws_full.append(headers)
    for row in data:
        ws_full.append([row.get(h, '') for h in headers])
    apply_colors(ws_full)

    # Detect faction column
    faction_key = next((h for h in headers if h.lower() in {'faction','camp','race'}), None)
    if faction_key:
        from collections import defaultdict
        groups = defaultdict(list)
        for row in data:
            label = str(row.get(faction_key, 'Unknown') or 'Unknown').strip() or 'Unknown'
            groups[label].append(row)

        used_titles = {ws_full.title}
        for faction, rows in groups.items():
            ws = wb.create_sheet(title=safe_sheet_title(faction, used_titles))
            ws.append(headers)
            for row in rows:
                ws.append([row.get(h, '') for h in headers])
            apply_colors(ws)

    # Ensure output dir exists and overwrite safely
    os.makedirs(os.path.dirname(excel_path), exist_ok=True)
    tmp_path = excel_path + ".tmp"
    wb.save(tmp_path)
    os.replace(tmp_path, excel_path)

if __name__ == "__main__":
    export_json_to_excel("herotier_full.json", EXCEL_PATH)
